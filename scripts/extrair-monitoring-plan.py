# Extrai o Monitoring Plan (xlsx) para JSON, para virar seed da tela Indicadores.
#
# POR QUE PYTHON NUM REPOSITORIO JS: e openpyxl que sabe resolver intervalo
# mesclado do Excel, e o plano depende disso - a planilha agrupa varios
# indicadores sob uma mesma atividade mesclando as colunas B a G, entao a
# atividade do indicador da linha 8 so existe na celula da linha 7. Ler sem
# resolver merge produz indicador orfao, que foi exatamente o que aconteceu na
# primeira extracao.
#
# RODAR:
#   python scripts/extrair-monitoring-plan.py "<caminho>/Monitoring Plan - EN.xlsx"
#
# Escreve docs/indicadores/monitoring-plan.json ao lado do repositorio.
#
# O QUE A PLANILHA E, em uma frase: o Plano de Monitoramento do projeto REDD+
# Awaete, com tres planos (Clima, Comunidade, Biodiversidade) derivados de uma
# Teoria da Mudanca, e para o plano de Comunidade tambem os valores ja medidos
# por periodo.
#
# DUAS ABAS DE COMUNIDADE, de proposito: ' Community MP' (sem espaco no fim) e a
# versao antiga, com os indicadores empilhados numa celula so como "i. ... ii.
# ...". ' Community MP ' (COM espaco no fim) e a evolucao feita pela propria
# equipe: um indicador por linha, mais as colunas de periodo. A segunda e a
# fonte; a primeira serve de conferencia da contagem.

import json
import re
import sys
from pathlib import Path

import openpyxl

# Nome exato das abas, com os espacos que existem de verdade. Nao "corrigir":
# ' Community MP' e ' Community MP ' sao duas abas diferentes, e a diferenca e
# um espaco no fim.
ABA_CLIMA = 'Climate MP'
ABA_COMUNIDADE_ANTIGA = ' Community MP'
ABA_COMUNIDADE = ' Community MP '
ABA_BIODIVERSIDADE = 'Biodiversity MP'
ABA_TOC = 'ToC'


def limpar(v):
    """Normaliza celula: espaco sobrando fora, quebra de linha tripla vira dupla."""
    if v is None:
        return None
    s = str(v).replace('\xa0', ' ').strip()
    if not s:
        return None
    s = re.sub(r'\n{3,}', '\n\n', s)
    s = re.sub(r'[ \t]+\n', '\n', s)
    # A planilha tem travessao em texto corrido; a regra do projeto proibe.
    # Escritos como escape de proposito: o caractere literal nao pode aparecer
    # nem no codigo que existe para elimina-lo.
    s = s.replace('\u2014', '-').replace('\u2013', '-')
    return s


def mapa_mesclado(planilha):
    """Devolve {(linha, coluna): (linha_ancora, coluna_ancora)} para cada celula
    dentro de um intervalo mesclado. O Excel guarda o valor so na ancora
    (canto superior esquerdo) e deixa o resto vazio."""
    m = {}
    for faixa in planilha.merged_cells.ranges:
        ancora = (faixa.min_row, faixa.min_col)
        for linha in range(faixa.min_row, faixa.max_row + 1):
            for coluna in range(faixa.min_col, faixa.max_col + 1):
                m[(linha, coluna)] = ancora
    return m


def ler(planilha, merges, linha, coluna):
    """Le a celula ja resolvendo merge."""
    ancora = merges.get((linha, coluna), (linha, coluna))
    return limpar(planilha.cell(row=ancora[0], column=ancora[1]).value)


def achar_cabecalho(planilha, merges, ate=12):
    """A linha de cabecalho nao e sempre a mesma: ha titulo e linhas em branco
    antes, e a quantidade varia por aba. Procura a linha que tem 'Activity' na
    coluna B."""
    for linha in range(1, ate + 1):
        if ler(planilha, merges, linha, 2) == 'Activity':
            return linha
    raise SystemExit(f'Nao achei a linha de cabecalho em {planilha.title!r}.')


def colunas(planilha, merges, linha_cab):
    """{titulo: indice} do cabecalho."""
    fora = {}
    for c in range(2, planilha.max_column + 1):
        t = ler(planilha, merges, linha_cab, c)
        if t:
            fora[t] = c
    return fora


# Marcador de lista romana no inicio de linha: "i.", "ii.", "iv.", e tambem
# "ii.%" (sem espaco), que aparece de verdade na coluna de unidade.
ROMANO = re.compile(r'(?m)^\s*(i{1,3}|iv|vi{1,3}|v|ix|x)\s*\.\s*')


def dividir_lista(texto):
    """Uma celula com "i. A / ii. B / iii. C" vira ['A', 'B', 'C'].

    Celula sem marcador volta como lista de um item so, para quem chama nao
    precisar distinguir os dois casos.

    ISTO NAO E COSMETICO: nas abas de Clima e Biodiversidade varios indicadores
    dividem uma linha, cada um com a SUA unidade, na mesma ordem. Guardar a
    celula inteira como um indicador so produziria um registro cuja unidade e
    'i. N/A ii. Hours/Day iii. N/A' - impossivel de somar, de comparar ou de
    colocar num eixo de grafico.
    """
    if not texto:
        return []
    partes = ROMANO.split(texto)
    if len(partes) == 1:
        return [texto.strip()]
    # split com grupo de captura intercala [antes, marcador, conteudo, ...].
    # O 'antes' costuma ser vazio; se vier com texto, e preambulo e se perde de
    # proposito - nunca foi indicador.
    itens = [partes[i + 1].strip() for i in range(1, len(partes) - 1, 2)]
    return [x for x in itens if x]


def extrair_plano(planilha, nome_plano):
    merges = mapa_mesclado(planilha)
    cab = achar_cabecalho(planilha, merges)
    cols = colunas(planilha, merges, cab)

    # As colunas de periodo sao tudo que vem depois de 'Frequency'. Nao da para
    # fixar a lista: a planilha ganha uma coluna a cada trimestre novo, e um
    # extrator que soubesse so ate '3th Quarter 2026' silenciosamente perderia
    # o 4o trimestre. Detectar por posicao continua funcionando.
    fim_estrutura = cols.get('Frequency')
    periodos = (
        {t: c for t, c in cols.items() if fim_estrutura and c > fim_estrutura}
        if fim_estrutura
        else {}
    )

    campos = {
        'atividade': 'Activity',
        'atividade_descricao': 'Activity Description',
        'output': 'Output',
        'outcome': 'Outcome',
        'impacto': 'Impact',
        'recurso': 'Resource',
        # So a aba de Clima tem esta coluna, e sem ela o indicador fica
        # ilegivel: o "indicador" ali e o simbolo da metodologia VM0048 ('AJ',
        # 'DLF', 'APA-Udef') e o significado ('Area of the jurisdiction') mora
        # aqui. Nas abas de Comunidade e Biodiversidade a coluna nao existe e o
        # proprio nome do indicador ja e a frase inteira.
        'descricao': 'Description',
        'unidade': 'Unit',
        'frequencia': 'Frequency',
    }
    # 'Indicators' no plural nas abas antigas, 'Indicator' no singular na nova.
    col_indicador = cols.get('Indicator') or cols.get('Indicators')

    linhas = []
    for r in range(cab + 1, planilha.max_row + 1):
        indicador = ler(planilha, merges, r, col_indicador) if col_indicador else None
        if not indicador:
            continue

        comum = {}
        for chave, titulo in campos.items():
            if titulo in cols:
                comum[chave] = ler(planilha, merges, r, cols[titulo])

        valores = {}
        for titulo, c in sorted(periodos.items(), key=lambda kv: kv[1]):
            v = ler(planilha, merges, r, c)
            # '-' e o jeito da equipe dizer "nao medido", diferente de zero.
            # Guardar como null preserva a distincao; virar 0 inventaria dado.
            if v is not None and v != '-':
                valores[titulo] = v

        nomes = dividir_lista(indicador)
        unidades = dividir_lista(comum.get('unidade'))
        frequencias = dividir_lista(comum.get('frequencia'))

        for i, nome in enumerate(nomes):
            reg = dict(comum)
            reg.update({'plano': nome_plano, 'linha_planilha': r, 'indicador': nome})

            # Unidade e frequencia pareiam POR POSICAO com o indicador. Quando a
            # celula tem um valor so e o indicador tem varios, o valor vale para
            # todos - e assim que a planilha usa 'Annual' para uma linha inteira.
            if unidades:
                reg['unidade'] = unidades[i] if i < len(unidades) else (
                    unidades[0] if len(unidades) == 1 else None
                )
            if frequencias:
                reg['frequencia'] = frequencias[i] if i < len(frequencias) else (
                    frequencias[0] if len(frequencias) == 1 else None
                )

            # 'N/A' na unidade quer dizer contagem simples (numero de cursos, de
            # pessoas). Guardar a string 'N/A' faria a tela imprimir "12 N/A".
            if reg.get('unidade') in ('N/A', 'n/a', 'N/a'):
                reg['unidade'] = None
            if reg.get('frequencia') == 'N/A':
                reg['frequencia'] = None

            # So o primeiro indicador da linha herda os valores medidos: as
            # colunas de periodo sao da LINHA, e quando a linha tem varios
            # indicadores nao da para saber de qual e o numero. Atribuir a todos
            # duplicaria a medicao. Na pratica isto nao perde nada, porque so a
            # aba de Comunidade tem valores e la ja e um indicador por linha.
            if valores and i == 0:
                reg['valores'] = valores
            if len(nomes) > 1:
                reg['compartilha_linha'] = True

            linhas.append(reg)

    return linhas


def extrair_toc(planilha):
    merges = mapa_mesclado(planilha)
    cab = None
    for linha in range(1, 12):
        if ler(planilha, merges, linha, 2) == 'STRATEGIC LINES':
            cab = linha
            break
    if cab is None:
        return []

    cols = colunas(planilha, merges, cab)
    campos = {
        'linha_estrategica': 'STRATEGIC LINES',
        'atividade': 'ACTIVITIES',
        'output': 'OUTPUTS',
        'outcome': 'OUTCOMES',
        'impacto': 'IMPACTS',
    }

    fora = []
    for r in range(cab + 1, planilha.max_row + 1):
        reg = {k: ler(planilha, merges, r, cols[t]) for k, t in campos.items() if t in cols}
        if reg.get('atividade'):
            fora.append(reg)
    return fora


def main():
    if len(sys.argv) < 2:
        raise SystemExit('Uso: python scripts/extrair-monitoring-plan.py "<arquivo.xlsx>"')

    caminho = Path(sys.argv[1])
    if not caminho.exists():
        raise SystemExit(f'Arquivo nao encontrado: {caminho}')

    livro = openpyxl.load_workbook(caminho, data_only=True)

    resultado = {
        'origem': caminho.name,
        'toc': extrair_toc(livro[ABA_TOC]) if ABA_TOC in livro.sheetnames else [],
        'planos': {},
    }

    for aba, nome in (
        (ABA_CLIMA, 'clima'),
        (ABA_COMUNIDADE, 'comunidade'),
        (ABA_BIODIVERSIDADE, 'biodiversidade'),
    ):
        if aba in livro.sheetnames:
            resultado['planos'][nome] = extrair_plano(livro[aba], nome)

    # Conferencia: a aba antiga de comunidade lista os indicadores empilhados
    # numa celula. Contar os "i. ii. iii." dela e comparar com a contagem da aba
    # nova diz se a nova esta completa ou se alguem parou de atualizar no meio.
    if ABA_COMUNIDADE_ANTIGA in livro.sheetnames:
        antiga = livro[ABA_COMUNIDADE_ANTIGA]
        merges = mapa_mesclado(antiga)
        cab = achar_cabecalho(antiga, merges)
        cols = colunas(antiga, merges, cab)
        ci = cols.get('Indicators') or cols.get('Indicator')
        total = 0
        if ci:
            for r in range(cab + 1, antiga.max_row + 1):
                texto = ler(antiga, merges, r, ci)
                if texto:
                    total += len(re.findall(r'(?m)^\s*(?:i{1,3}|iv|v|vi{1,3}|ix|x)\s*\.', texto))
        resultado['conferencia'] = {
            'comunidade_aba_antiga_indicadores': total,
            'comunidade_aba_nova_indicadores': len(resultado['planos'].get('comunidade', [])),
        }

    destino = Path(__file__).resolve().parents[1] / 'docs' / 'indicadores'
    destino.mkdir(parents=True, exist_ok=True)
    arquivo = destino / 'monitoring-plan.json'
    arquivo.write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding='utf-8')

    print(f'Escrito: {arquivo}')
    for nome, linhas in resultado['planos'].items():
        com_valor = sum(1 for x in linhas if x.get('valores'))
        print(f'  {nome:16s} {len(linhas):3d} indicadores, {com_valor:3d} com valor medido')
    print(f'  {"toc":16s} {len(resultado["toc"]):3d} atividades')
    if 'conferencia' in resultado:
        c = resultado['conferencia']
        print(
            f'  conferencia: aba antiga {c["comunidade_aba_antiga_indicadores"]}, '
            f'aba nova {c["comunidade_aba_nova_indicadores"]}'
        )


main()
