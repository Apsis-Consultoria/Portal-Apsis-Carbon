# -*- coding: utf-8 -*-
"""
Gera o seed COMPLETO de prestacao de contas e atividades de campo a partir das
TRES planilhas da operacao. Versao 3 (01/09/2026).

    cd portal-apsis-carbon; python scripts/gerar-seed-prestacao.py "C:/Users/<voce>/Downloads"

-------------------------------------------------------------------------------
O QUE MUDOU DA V2, e por que (cada item nasceu de releitura da planilha)
-------------------------------------------------------------------------------
1. DESCRICAO DO GRUPO DE BAIXO ENTRA INTEIRA. A releitura provou que as 10
   descricoes distintas que o vocabulario retinha ("1 Cota do Boi", "2 Paineis
   solares", "Repasse direto a Associacao...") sao itens, nao pessoas. So a
   ajuda de custo carrega nome, e continua retida.
2. CARGO FICA, NOME SAI. No ciclo 2 do Baixo as linhas nominais tem o padrao
   "Cargo - Nome" (Presidente, Coordenador, 1o Tesoureiro). O cargo e dado
   organizacional e fica; o nome e retido.
3. TEM_COMPROVANTE DERIVADO DAS OBSERVACOES do Cima. O analista escreveu a
   conciliacao na coluna de observacao ("Nao tem qualquer recibo/comprovante",
   "Ok.") e a v2 jogava isso fora - por isso o Cima aparecia 100% sem
   comprovante. Agora o padrao textual vira o campo.
4. OBSERVACAO ENTRA POR FREQUENCIA. Frase que se repete 3+ vezes e status de
   conciliacao ("RECIBO DUPLICADO", "Nao esta na prestacao") e entra; frase
   unica e onde mora o nome de pessoa, e fica de fora. Vale para as abas de
   periodo e para OBS 1/OBS 2 da base de comprovantes.
5. ATIVIDADES DE CAMPO: as abas MR - 1, MR -2, Set - Dez 25 e MR 3 - 2026 de
   "Atividade Parakana.xlsx" entram em carbon_atividades_campo. A coluna
   Responsavel (nome) NAO entra; os textos passam pelo vocabulario de
   pseudonimizacao das atas (termos com sinal_pessoa=1 viram o codigo [Pn]).
6. TRAVESSAO E MEIA-RISCA normalizados para hifen em todo texto importado.

CONTINUA DE FORA por decisao: abas de rascunho/motor, "Pendencias*" (nominal;
a pendencia real ja esta em tem_comprovante=false), coluna Nome/CPF de
comprovantes (o documento fisico e achavel por mes + ordem no mes).
"""

import csv
import io
import re
import sys
import unicodedata
from calendar import monthrange
from collections import Counter
from datetime import date, datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit('Falta openpyxl. Rode: pip install openpyxl')

AQUI = Path(__file__).resolve().parent
RAIZ = AQUI.parent
SAIDA = RAIZ / 'supabase' / 'seeds' / 'prestacao_contas.sql'
PASTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads'

ARQ_BAIXO = PASTA / 'Antecipação Grupo de Baixo.xlsx'
ARQ_CIMA = PASTA / 'Antecipação Grupo de Cima.xlsx'
ARQ_ATIV = PASTA / 'Atividade Parakanã.xlsx'
ARQ_TERMOS = Path('C:/Users/FilipeOliveiraAPSISC/notion-export/revisao-termos-atas.csv')
# Lista revisada de nome de pessoa das planilhas da prestacao. FORA do repositorio
# porque ela e a propria lista de nomes. Sem ela o script aborta (ver
# carregar_termos_pessoa). Colunas: codigo;termo;motivo
ARQ_NOMES = Path('C:/Users/FilipeOliveiraAPSISC/notion-export/nomes-prestacao.csv')

PADRAO_AJUDA = re.compile(r'ajuda de custo', re.I)
# "Presidente - Fulano", "1o Tesoureiro - Fulano": o cargo fica, o nome sai.
PADRAO_CARGO = re.compile(
    r'^\s*((?:1[oº°]?|2[oº°]?)?\s*(?:presidente|vice[- ]?presidente|coordenador[a]?'
    r'|tesoureir[oa]|secretari[oa]|diretor[a]?))\s*[-\u2013\u2014]\s*\S',
    re.I,
)

# Vocabulario de despesa (so para o Grupo de CIMA, cuja coluna mistura item com
# nome de pessoa; no Baixo a releitura provou que a coluna e so item).
VOCAB_DESPESA = re.compile(
    r'tarifa|mensalidade|despesa|assessoria|jur[ií]dic|advogad|contador|contabil'
    r'|diretoria|sal[aá]rio|gasolina|combust|diesel|[oó]leo|cesta|dep[oó]sito'
    r'|internet|starlink|energia|placa|bateria|inversor|controlador|motor|gerador'
    r'|frete|passag|transporte|barco|voadeira|rabeta|motosserra|po[çc]o|bomba'
    r'|constru|telha|cimento|ferramenta|material|servi[çc]|manuten|aliment'
    r'|refei|almo[çc]|marmita|lanche|hospedagem|viagem|aluguel|imposto|taxa'
    r'|multa|cart[óo]rio|documenta|caixa|banco|anuidade|seguro|farinha|semente'
    r'|muda|ro[çc]a|a[çc]a[ií]|castanha|equipamento|pe[çc]a|reparo|instala'
    r'|perfura|antena|celular|notebook|impressora|freezer|geladeira|fog[aã]o'
    r'|botij|g[aá]s|rancho|compra|pagamento de conta|certificado|cota|boi'
    r'|painel|paine|repasse|transfer[eê]ncia|carro|camionete|honor[aá]rio',
    re.I,
)

# Conciliacao escrita pelo analista na observacao -> tem_comprovante.
OBS_SEM_COMPROVANTE = re.compile(
    r'n[aã]o (?:tem|h[aá]|encontrad|localizad|identificad|foi poss[ií]vel)|sem comprovante', re.I)
OBS_COM_COMPROVANTE = re.compile(r'^\s*ok\b|h[aá] comprovante', re.I)

RETIDA_SALARIO = 'Salário de diretoria/equipe da associação (nome retido - LGPD)'
RETIDA_PADRAO = 'Repasse a morador (identificação no comprovante físico - LGPD)'
RETIDA_AJUDA = 'Ajuda de custo (nome retido - LGPD)'


def sql(v):
    if v is None:
        return 'null'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return repr(round(v, 2))
    if isinstance(v, (date, datetime)):
        return "'%s'" % v.strftime('%Y-%m-%d')
    return "'%s'" % str(v).replace("'", "''")


def texto(v):
    if v is None:
        return None
    # Travessao e meia-risca nunca entram no banco (regra 3 do CLAUDE.md).
    s = str(v).replace('\u2014', '-').replace('\u2013', '-').strip()
    return s or None


def numero(v):
    if v in (None, ''):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('R$', '').replace(' ', '')
    if s in ('-', ''):
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def dia(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def ultimo_dia(d):
    if d is None:
        return None
    return date(d.year, d.month, monthrange(d.year, d.month)[1])


def chave(nome):
    s = unicodedata.normalize('NFKD', str(nome or '')).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', '', s.lower())


# ===== Pseudonimizacao ========================================================
# Termo que e nome de pessoa vira um marcador estavel [Pn], o mesmo marcador para a
# mesma pessoa em todos os seeds. E a convencao ja usada em seeds/atas_reunioes.sql.
#
# -----------------------------------------------------------------------------
# ESTA FUNCAO JA DEIXOU NOME VAZAR. Duas causas, as duas corrigidas em 02/09/2026:
# -----------------------------------------------------------------------------
# 1. `sinal_pessoa` E CONTAGEM, NAO SINALIZADOR. O teste era `linha[3] == '1'`,
#    comparacao de string com uma coluna que guarda QUANTAS vezes a heuristica viu
#    sinal de pessoa. Joedson tem 5, Marcelo 12, Tyge 3: nenhum casava, e os tres
#    entraram no seed com o nome cru. Sairam 19 substituicoes onde as atas fizeram
#    6029.
#
#    O CRITERIO CONTINUA `== '1'` DE PROPOSITO, e nao foi alargado para `> 0`.
#    Medido: dos 184 termos com sinal_pessoa > 0, 24 aparecem nestes seeds, e so
#    TRES sao pessoa (Joedson, Marcelo, Tyge). Os outros 21 sao ruido da heuristica
#    - "Nao" tem 7, "Alta" 12, "Falta" 9 - e um deles, `Xeteria` (sinal 2), e ALDEIA.
#    Alargar trocaria "Nao tem comprovante" por "[P862] tem comprovante" e apagaria
#    o nome de uma aldeia, que e justamente a dimensao que as telas de rateio usam.
#    Quem entra vem da lista revisada de ARQ_NOMES, abaixo.
#
# 2. A DESCRICAO DO LANCAMENTO NAO PASSAVA AQUI. `pseudonimizar` era aplicada so em
#    atividade, evidencia e obs. A coluna de descricao do lancamento e onde moram
#    "Combustivel Konomiria" e "Suporte alimentacao comunidade - Mama": o
#    VOCAB_DESPESA casava em "combust", concluia "isto e item, nao e nome" e
#    devolvia o texto intacto COM o nome. A celula tem as duas coisas.
#
# ARQ_NOMES FICA FORA DO REPOSITORIO porque ele E a lista de nomes: versiona-lo
# desfaria a pseudonimizacao pela porta dos fundos. Mesma decisao do
# revisao-termos-atas.csv. Sem ele, este script ABORTA: o padrao anterior era
# devolver mapa vazio e gerar o seed com os nomes crus, calado.
def carregar_termos_pessoa():
    mapa = {}

    # 1. Os que a heuristica das atas marcou com sinal exatamente 1. Ver acima por
    #    que este criterio nao foi alargado.
    if ARQ_TERMOS.exists():
        with io.open(ARQ_TERMOS, encoding='utf-8-sig') as f:
            for linha in csv.reader(f, delimiter=';'):
                if len(linha) > 3 and linha[3] == '1' and linha[1]:
                    mapa[linha[1]] = linha[0]

    # 2. A lista revisada termo a termo para ESTAS planilhas. Autoritativa: sobrepoe
    #    a heuristica, que errou a classificacao do Damiao (sinal_org=1, e ele e
    #    pessoa) e nao conhece quem so aparece aqui (Konomiria, Moipa, Kaoe...).
    if not ARQ_NOMES.exists():
        sys.exit(
            'ABORTADO: nao encontrei %s.\n'
            'Esse arquivo carrega os nomes de pessoa das planilhas da prestacao e\n'
            'vive FORA do repositorio de proposito. Sem ele o seed sairia com nome\n'
            'de pessoa em texto livre, ligado a valor gasto, o que e dado pessoal\n'
            'sob a LGPD e nao pode entrar no git.' % ARQ_NOMES
        )
    #    Codigo MANTER e VETO: tira o termo do mapa. Existe porque a heuristica das
    #    atas marca palavra comum como pessoa - "Tem", "Houve", "Carro" - e marcou
    #    MAROXEWARA, que e ALDEIA, com sinal_pessoa=1. Enquanto pseudonimizar()
    #    rodava so nas atividades isso quase nao aparecia; aplicada na descricao do
    #    lancamento, "Tem foto?" virou "[P875] foto?".
    with io.open(ARQ_NOMES, encoding='utf-8-sig') as f:
        r = csv.reader(f, delimiter=';')
        next(r, None)  # cabecalho
        for linha in r:
            if len(linha) < 2 or not linha[0].strip() or not linha[1].strip():
                continue
            codigo, termo = linha[0].strip(), linha[1].strip()
            if codigo == 'MANTER':
                mapa.pop(termo, None)
            else:
                mapa[termo] = codigo
    return mapa


TERMOS_PESSOA = carregar_termos_pessoa()
# Mais longo primeiro: "Flavio Rodrigues Sousa" tem que casar antes de "Flavio",
# senao o nome completo sai como tres marcadores em sequencia.
_padrao_termos = re.compile(
    r'\b(' + '|'.join(re.escape(t) for t in sorted(TERMOS_PESSOA, key=len, reverse=True)) + r')\b')

trocas_pessoa = Counter()


def pseudonimizar(t):
    if not t:
        return t

    def troca(m):
        trocas_pessoa[m.group(1)] += 1
        return TERMOS_PESSOA[m.group(1)]

    return _padrao_termos.sub(troca, t)


# ===== Estado ================================================================

CICLOS = {
    ('baixo', 'c1'): {'nome': 'Outubro 2024 a Abril 2025', 'inicio': date(2024, 10, 1), 'fim': date(2025, 4, 30), 'abertura': None},
    ('baixo', 'c2'): {'nome': 'Maio a Julho 2025', 'inicio': date(2025, 5, 1), 'fim': date(2025, 7, 31), 'abertura': None},
    ('cima', 'c1'): {'nome': 'Outubro 2024 a Abril 2025', 'inicio': date(2024, 10, 1), 'fim': date(2025, 4, 30), 'abertura': None},
    ('cima', 'c2'): {'nome': 'Maio a Setembro 2025', 'inicio': date(2025, 5, 1), 'fim': date(2025, 9, 30), 'abertura': None},
}

lancamentos = []
antecipacoes = []
comprovantes = []
atividades = []
retidos = Counter()
descartes = []
aldeias = {}
eixos = {}


def registrar_aldeia(grupo, nome):
    n = texto(nome)
    if not n:
        return None
    k = chave(n)
    if not k:
        return None
    if k not in aldeias:
        aldeias[k] = (grupo, n)
    return k


def registrar_eixo(nome):
    n = texto(nome)
    if not n:
        return None
    k = chave(n)
    if not k:
        return None
    if k not in eixos:
        eixos[k] = n
    return k


def descricao_baixo(bruta, linha, aba):
    """No Baixo a coluna e de itens; so ajuda de custo e cargo-nome retem.

    A releitura de 02/09/2026 mostrou que "coluna e de itens" nao era verdade
    inteira: "Suporte alimentacao comunidade - Mama" esta nela, e o beneficiario
    depois do hifen as vezes e aldeia e as vezes e pessoa. Por isso o texto que
    sobrevive ainda passa por pseudonimizar().
    """
    t = texto(bruta)
    if not t:
        return 'Lancamento %s linha %d' % (aba, linha), False
    if PADRAO_AJUDA.search(t):
        retidos['%s: ajuda de custo' % aba] += 1
        return RETIDA_AJUDA, True
    m = PADRAO_CARGO.match(t)
    if m:
        retidos['%s: cargo - nome' % aba] += 1
        return '%s da associação (nome retido - LGPD)' % m.group(1).strip().title(), True
    return pseudonimizar(t), False


def descricao_cima(bruta, eixo_nome, linha, aba):
    """No Cima a coluna mistura item com nome de pessoa: vocabulario decide.

    ATENCAO ao caminho do VOCAB_DESPESA: ele responde "isto e item, nao e nome" e
    devolvia o texto INTACTO. Mas a celula tem as duas coisas ao mesmo tempo -
    "Combustivel Konomiria" casa em `combust` e carregava o nome para o banco.
    O item fica, o nome vira marcador.
    """
    t = texto(bruta)
    if not t:
        return 'Lancamento %s linha %d' % (aba, linha), False
    if PADRAO_AJUDA.search(t):
        retidos['%s: ajuda de custo' % aba] += 1
        return RETIDA_AJUDA, True
    if VOCAB_DESPESA.search(t):
        return pseudonimizar(t), False
    retidos['%s: nominal' % aba] += 1
    if eixo_nome and 'salari' in chave(eixo_nome):
        return RETIDA_SALARIO, True
    return RETIDA_PADRAO, True


def comprovante_da_obs(obs):
    """A conciliacao que o analista escreveu vira o campo tem_comprovante."""
    t = texto(obs)
    if not t:
        return None
    if OBS_SEM_COMPROVANTE.search(t):
        return False
    if OBS_COM_COMPROVANTE.search(t):
        return True
    return None


# ===== Grupo de Baixo ========================================================

def ler_baixo():
    wb = load_workbook(ARQ_BAIXO, data_only=True)

    ws = wb['Extrato']  # cabecalho L4, dados L5+
    for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True), 5):
        if not any(c not in (None, '') for c in row):
            continue
        tipo = texto(row[1])
        valor = numero(row[7])
        competencia = dia(row[5]) or dia(row[4])
        if not tipo or competencia is None:
            continue
        if valor is None:
            descartes.append(('Baixo/Extrato L%d' % i, 'linha sem valor (receita de abril, divergencia ja aberta)'))
            continue
        if tipo.lower().startswith('receita'):
            antecipacoes.append({'grupo': 'baixo', 'ciclo': 'c1', 'competencia': ultimo_dia(competencia),
                                 'valor': abs(valor), 'aba': 'Extrato', 'linha': i})
            continue
        descricao, _ = descricao_baixo(row[3], i, 'Extrato')
        comprov = texto(row[10])
        lancamentos.append({
            'grupo': 'baixo', 'ciclo': 'c1', 'competencia': ultimo_dia(competencia),
            'descricao': descricao,
            'valor': -abs(valor) if valor > 0 else valor,
            'quantidade': numero(row[2]),
            'aldeia': registrar_aldeia('baixo', row[6]),
            'eixo': registrar_eixo(row[12]),
            'documento': texto(row[9]),
            'tem_comprovante': None if not comprov or comprov.upper() == 'NA' else comprov.lower().startswith('s'),
            'obs_bruta': texto(row[14] if len(row) > 14 else None),
            'aba': 'Extrato', 'linha': i,
        })

    ws = wb['Maio - Julho 2025']  # cabecalho L3; SO colunas 1..9 (11+ e rascunho lateral)
    for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True), 4):
        if not any(c not in (None, '') for c in row[:10]):
            continue
        tipo = texto(row[1])
        if not tipo:
            continue
        valor = numero(row[6])
        if tipo.lower().startswith('saldo'):
            if valor is not None:
                CICLOS[('baixo', 'c2')]['abertura'] = valor
            continue
        if valor is None:
            continue
        if tipo.lower().startswith('receita'):
            comp = ultimo_dia(dia(row[4])) or date(2025, 7, 31)
            antecipacoes.append({'grupo': 'baixo', 'ciclo': 'c2', 'competencia': comp,
                                 'valor': abs(valor), 'aba': 'Maio - Julho 2025', 'linha': i})
            continue
        descricao, _ = descricao_baixo(row[3], i, 'Maio - Julho 2025')
        # A coluna Mes deste bloco e o TEXTO "Maio, Junho e Julho - 25": sem mes
        # por linha, a competencia vai para o fim do ciclo, rastreavel na origem.
        lancamentos.append({
            'grupo': 'baixo', 'ciclo': 'c2', 'competencia': date(2025, 7, 31),
            'descricao': descricao,
            'valor': -abs(valor) if valor > 0 else valor,
            'quantidade': numero(row[2]),
            'aldeia': registrar_aldeia('baixo', row[5]),
            'eixo': registrar_eixo(row[9]),
            'documento': None, 'tem_comprovante': None, 'obs_bruta': None,
            'aba': 'Maio - Julho 2025', 'linha': i,
        })


# ===== Grupo de Cima =========================================================

def ler_cima():
    wb = load_workbook(ARQ_CIMA, data_only=True)

    # Antecipacoes: aba Resumo, L11 datas / L12 valores.
    datas = [c.value for c in wb['Resumo'][11]]
    valores = [c.value for c in wb['Resumo'][12]]
    for d, v in zip(datas, valores):
        dd, vv = dia(d), numero(v)
        if dd is None or vv is None or vv <= 0:
            continue
        antecipacoes.append({'grupo': 'cima', 'ciclo': 'c1' if dd <= date(2025, 4, 30) else 'c2',
                             'competencia': ultimo_dia(dd), 'valor': vv, 'aba': 'Resumo', 'linha': 12})

    # Despesas: as duas abas de periodo. Cols: 3 Nome/Item, 4 Valor, 5 Obs,
    # 6 Aldeia, 7 Mes(data), 8 Eixo.
    for aba, primeira, ciclo in [('Out 24 a Abril 25', 4, 'c1'), ('Maio a setembro 25', 3, 'c2')]:
        ws = wb[aba]
        for i, row in enumerate(ws.iter_rows(min_row=primeira, values_only=True), primeira):
            valor = numero(row[4]) if len(row) > 4 else None
            comp = dia(row[7]) if len(row) > 7 else None
            if valor is None or comp is None:
                continue
            eixo_nome = texto(row[8]) if len(row) > 8 else None
            descricao, _ = descricao_cima(row[3] if len(row) > 3 else None, eixo_nome, i, aba)
            obs = texto(row[5]) if len(row) > 5 else None
            lancamentos.append({
                'grupo': 'cima', 'ciclo': ciclo, 'competencia': ultimo_dia(comp),
                'descricao': descricao,
                'valor': -abs(valor) if valor > 0 else valor,
                'quantidade': None,
                'aldeia': registrar_aldeia('cima', row[6] if len(row) > 6 else None),
                'eixo': registrar_eixo(eixo_nome),
                'documento': None,
                # A conciliacao do analista, escrita na observacao, vira o campo.
                'tem_comprovante': comprovante_da_obs(obs),
                'obs_bruta': obs,
                'aba': aba, 'linha': i,
            })

    # Comprovantes: Base de dados, cabecalho L4. Nome (2) e CPF (3) NAO entram.
    ws = wb['Base de dados']
    for i, row in enumerate(ws.iter_rows(min_row=5, values_only=True), 5):
        if not any(c not in (None, '') for c in row):
            continue
        valor = numero(row[8])
        quando = dia(row[9]) or dia(row[10])
        if valor is None or valor <= 0 or quando is None:
            descartes.append(('Cima/Base de dados L%d' % i, 'sem valor positivo ou sem data'))
            continue
        obs = ' / '.join(t for t in (texto(row[12] if len(row) > 12 else None),
                                     texto(row[13] if len(row) > 13 else None)) if t) or None
        comprovantes.append({
            'grupo': 'cima', 'ciclo': 'c1' if quando <= date(2025, 4, 30) else 'c2',
            'ordem': int(numero(row[1])) if numero(row[1]) is not None else None,
            'data': quando, 'valor': valor,
            'inst_recebedor': texto(row[4]), 'inst_pagador': texto(row[7]),
            'obs_bruta': obs,
            'aba': 'Base de dados', 'linha': i,
        })


# ===== Atividades de campo ===================================================

STATUS_ATIV = {'concluido': 'Concluído', 'ok': 'Concluído', 'duvida': 'Dúvida',
               'pendente': 'Pendente', 'andamento': 'Em andamento'}


def status_ativ(v):
    t = texto(v)
    if not t:
        return None
    return STATUS_ATIV.get(chave(t), t)


def ler_atividades():
    wb = load_workbook(ARQ_ATIV, data_only=True)

    # MR - 1: 1 Dia, 2 Mes, 3 Ano, 4 Data, 6 Atividade, 7 LE1, 8 LE2, 9 Grupo,
    # 10 Valor, 11 Evidencias, 13 Status. Colunas 12 Link, 14 Comentarios e
    # 15 Responsavel NAO entram (a ultima e nome de pessoa).
    ws = wb['MR - 1 ']
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        ativ = texto(row[6]) if len(row) > 6 else None
        if not ativ:
            continue
        quando = dia(row[4])
        if quando is None:
            try:
                quando = date(int(row[3]), int(row[2]), int(row[1]))
            except (TypeError, ValueError):
                quando = None
        grupo = chave(row[9] if len(row) > 9 else None)
        atividades.append({
            'relatorio': 'MR-1', 'inicio': quando, 'termino': quando,
            'atividade': pseudonimizar(ativ),
            'instituicao': None,
            'tipo': None,
            'linha': ' / '.join(t for t in (texto(row[7]), texto(row[8])) if t) or None,
            'evidencia': pseudonimizar(texto(row[11]) if len(row) > 11 else None),
            'valor': numero(row[10]) if len(row) > 10 else None,
            'status': status_ativ(row[13] if len(row) > 13 else None),
            'obs': None,
            'grupo': 'cima' if 'cima' in grupo or 'alto' in grupo else ('baixo' if 'baixo' in grupo else None),
            'aba': 'MR - 1', 'linha_origem': i,
        })

    # As tres abas de periodo tem layouts DIFERENTES entre si; um mapa de
    # indices explicito por aba e o que impede a coluna certa de escorregar.
    #   MR -2        1 Mes, 2 Inicio, 3 Termino, 4 Inst, 5 Tipo, 6 Ativ, 7 Evid, 8 Status, 9 Obs
    #   Set - Dez 25 0 Mes, 1 Inicio, 2 Termino, 3 Inst, 4 Tipo, 5 Ativ, 6 Evid, 7 Status, 8 Obs
    #   MR 3 - 2026        1 Inicio, 2 Termino, 3 Inst, 4 Tipo, 5 Ativ, 6 Evid, 7 Status, 8 Obs
    MAPAS = [
        ('MR -2 ', 'MR-2', {'ini': 2, 'fim': 3, 'inst': 4, 'tipo': 5, 'ativ': 6, 'evid': 7, 'st': 8, 'obs': 9}),
        ('Set - Dez 25', 'MR-3', {'ini': 1, 'fim': 2, 'inst': 3, 'tipo': 4, 'ativ': 5, 'evid': 6, 'st': 7, 'obs': 8}),
        ('MR 3 - 2026', 'MR-3', {'ini': 1, 'fim': 2, 'inst': 3, 'tipo': 4, 'ativ': 5, 'evid': 6, 'st': 7, 'obs': 8}),
    ]
    for aba, rel, m in MAPAS:
        ws = wb[aba]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            def col(nome):
                j = m[nome]
                return row[j] if j < len(row) else None
            ativ = texto(col('ativ'))
            if not ativ:
                continue
            atividades.append({
                'relatorio': rel, 'inicio': dia(col('ini')), 'termino': dia(col('fim')),
                'atividade': pseudonimizar(ativ),
                'instituicao': texto(col('inst')), 'tipo': texto(col('tipo')), 'linha': None,
                'evidencia': pseudonimizar(texto(col('evid'))),
                'valor': None, 'status': status_ativ(col('st')),
                'obs': pseudonimizar(texto(col('obs'))),
                'grupo': None, 'aba': aba.strip(), 'linha_origem': i,
            })


ler_baixo()
ler_cima()
ler_atividades()

# ===== Pos-processamento das observacoes =====================================
# Frase repetida 3+ vezes e status de conciliacao e entra; frase unica tende a ser
# anotacao de caso especifico, e fica de fora.
#
# ATENCAO: A FREQUENCIA NAO DIZ NADA SOBRE HAVER NOME, e o comentario que estava
# aqui afirmava que dizia ("frase unica e onde mora o nome de pessoa"). Estava
# errado, e o erro custou dois vazamentos:
#   'Recibo Assinado pelo Awyakynga.'   aparece 5 vezes
#   'Comprovante em nome de Myrytyga'   aparece 3 vezes
# As duas passaram o corte JUSTAMENTE por serem repetidas. Um procedimento
# repetido - "o recibo e sempre assinado pela mesma pessoa" - produz frase
# repetida COM nome.
#
# Por isso o corte de frequencia continua (ele reduz exposicao de anotacao
# solta), mas o que passa ainda vai para pseudonimizar().
freq = Counter()
for l in lancamentos:
    if l['obs_bruta']:
        freq[l['obs_bruta']] += 1
for c in comprovantes:
    if c['obs_bruta']:
        freq[c['obs_bruta']] += 1

obs_descartadas = 0
for item in lancamentos + comprovantes:
    bruta = item.pop('obs_bruta')
    if bruta and freq[bruta] >= 3:
        item['observacoes'] = pseudonimizar(bruta)
    else:
        item['observacoes'] = None
        if bruta:
            obs_descartadas += 1

# ===== Conciliacao automatica por valor e mes ================================
# E o que o analista fazia a mao: um comprovante de mesmo valor, no mesmo mes,
# do mesmo ciclo, lastreia um lancamento. Multiconjunto de proposito: tres
# lancamentos de 189 no mes so viram "com comprovante" se houver tres
# comprovantes de 189 no mes. So mexe em quem esta como "nao informado" - o que
# o analista marcou a mao (True ou False) prevalece sempre.
from collections import defaultdict
saldo_comp = defaultdict(int)
for cp in comprovantes:
    saldo_comp[(cp['ciclo'], cp['data'].strftime('%Y-%m'), round(cp['valor'], 2))] += 1

conciliados = 0
for l in lancamentos:
    if l['grupo'] != 'cima' or l['tem_comprovante'] is not None:
        continue
    k = (l['ciclo'], l['competencia'].strftime('%Y-%m'), round(abs(l['valor']), 2))
    if saldo_comp.get(k, 0) > 0:
        saldo_comp[k] -= 1
        l['tem_comprovante'] = True
        l['observacoes'] = ((l['observacoes'] + ' / ') if l['observacoes'] else '') +             'Conciliado automaticamente por valor e mes na importacao'
        conciliados += 1

# ===== Escrita ===============================================================

GRUPOS = [('baixo', 'Grupo de Baixo', ['baixo', 'grupo de baixo', 'parakana de baixo']),
          ('cima', 'Grupo de Cima', ['cima', 'alto', 'grupo de cima', 'parakana de cima'])]

L = []
w = L.append
w('-- Gerado por scripts/gerar-seed-prestacao.py (v3). Nao edite a mao.')
w('-- SEM DADO PESSOAL: ver o cabecalho do script.')
w('')
w('begin;')
w('')
w('do $$')
w('begin')
w("  if not exists (select 1 from public.carbon_projetos")
w("     where nome ilike '%parakan%' or nome ilike '%awaet%') then")
w("    raise exception 'nenhum projeto Parakana/Awaete em carbon_projetos';")
w('  end if;')
w('end $$;')
w('')
w('-- Limpeza da carga anterior: SO linha importada (origem_aba preenchida) e os')
w('-- ciclos da v1. Nada digitado pela tela e tocado.')
w('delete from public.carbon_comprovantes where origem_aba is not null;')
w('delete from public.carbon_prestacao_lancamentos where origem_aba is not null;')
w('delete from public.carbon_antecipacoes where origem_aba is not null;')
w("delete from public.carbon_ciclos_prestacao where nome like 'Antecipacao %';")
w('delete from public.carbon_atividades_campo where origem_aba is not null;')
w('')

for ch, nome, apel in GRUPOS:
    w("insert into public.carbon_grupos_comunitarios (projeto_id, chave, nome, apelidos)")
    w("select p.id, %s, %s, %s::text[] from public.carbon_projetos p" % (
        sql(ch), sql(nome), sql('{' + ','.join('"%s"' % a for a in apel) + '}')))
    w("  where p.nome ilike '%parakan%' or p.nome ilike '%awaet%'")
    w("on conflict (projeto_id, chave) do update set nome = excluded.nome, atualizado_em = now();")
w('')

for k, (grupo, nome) in sorted(aldeias.items()):
    w("insert into public.carbon_aldeias (grupo_id, nome, e_associacao)")
    w("select g.id, %s, %s from public.carbon_grupos_comunitarios g where g.chave = %s"
      % (sql(nome), sql('associac' in k), sql(grupo)))
    w("on conflict (grupo_id, nome) do nothing;")
w('')

for i, (k, nome) in enumerate(sorted(eixos.items())):
    w("insert into public.carbon_eixos (projeto_id, nome, ordem)")
    w("select p.id, %s, %d from public.carbon_projetos p" % (sql(nome), i * 10))
    w("  where p.nome ilike '%parakan%' or p.nome ilike '%awaet%'")
    w("on conflict (projeto_id, nome) do nothing;")
w('')

for (grupo, cc), c in CICLOS.items():
    w("insert into public.carbon_ciclos_prestacao (projeto_id, grupo_id, nome, inicio, fim, saldo_abertura, status)")
    w("select g.projeto_id, g.id, %s, %s, %s, %s, 'em_conciliacao'" % (
        sql(c['nome']), sql(c['inicio']), sql(c['fim']), sql(c['abertura'])))
    w("  from public.carbon_grupos_comunitarios g where g.chave = %s" % sql(grupo))
    w("on conflict (grupo_id, nome) do update set inicio = excluded.inicio,")
    w("  fim = excluded.fim, saldo_abertura = excluded.saldo_abertura, atualizado_em = now();")
w('')


def sel_ciclo(grupo, cc):
    return ("from public.carbon_ciclos_prestacao c"
            " join public.carbon_grupos_comunitarios g on g.id = c.grupo_id"
            " where g.chave = %s and c.nome = %s" % (sql(grupo), sql(CICLOS[(grupo, cc)]['nome'])))


for a in antecipacoes:
    w("insert into public.carbon_antecipacoes (ciclo_id, competencia, valor, origem_aba, origem_linha)")
    w("select c.id, %s, %s, %s, %d %s" % (
        sql(a['competencia']), sql(a['valor']), sql(a['aba']), a['linha'], sel_ciclo(a['grupo'], a['ciclo'])))
    w("on conflict (ciclo_id, competencia) do update set valor = excluded.valor;")
w('')

for l in lancamentos:
    w("insert into public.carbon_prestacao_lancamentos")
    w("  (ciclo_id, grupo_id, aldeia_id, eixo_id, competencia, descricao, valor,")
    w("   quantidade, documento, tem_comprovante, observacoes, origem_aba, origem_linha)")
    w("select c.id, g.id,")
    w("  (select a.id from public.carbon_aldeias a where a.grupo_id = g.id and a.nome = %s),"
      % sql(aldeias[l['aldeia']][1] if l['aldeia'] else None))
    w("  (select e.id from public.carbon_eixos e where e.projeto_id = g.projeto_id and e.nome = %s),"
      % sql(eixos[l['eixo']] if l['eixo'] else None))
    w("  %s, %s, %s, %s, %s, %s, %s, %s, %d" % (
        sql(l['competencia']), sql(l['descricao']), sql(l['valor']), sql(l['quantidade']),
        sql(l['documento']), sql(l['tem_comprovante']), sql(l['observacoes']),
        sql(l['aba']), l['linha']))
    w("  %s" % sel_ciclo(l['grupo'], l['ciclo']))
    w("on conflict on constraint carbon_prest_lanc_origem_uq do update set")
    w("  valor = excluded.valor, descricao = excluded.descricao,")
    w("  tem_comprovante = excluded.tem_comprovante, observacoes = excluded.observacoes,")
    w("  atualizado_em = now();")
w('')

for cp in comprovantes:
    w("insert into public.carbon_comprovantes")
    w("  (ciclo_id, grupo_id, aldeia_id, ordem_no_mes, data, valor,")
    w("   instituicao_recebedor, instituicao_pagador, observacoes, origem_aba, origem_linha)")
    w("select c.id, g.id,")
    w("  (select a.id from public.carbon_aldeias a where a.grupo_id = g.id and a.nome = %s),"
      % sql(aldeias[cp['aldeia']][1] if cp.get('aldeia') else None))
    w("  %s, %s, %s, %s, %s, %s, %s, %d" % (
        sql(cp['ordem']), sql(cp['data']), sql(cp['valor']),
        sql(cp['inst_recebedor']), sql(cp['inst_pagador']), sql(cp['observacoes']),
        sql(cp['aba']), cp['linha']))
    w("  %s" % sel_ciclo(cp['grupo'], cp['ciclo']))
    w("on conflict on constraint carbon_comprovantes_origem_uq do update set")
    w("  valor = excluded.valor, data = excluded.data, observacoes = excluded.observacoes,")
    w("  atualizado_em = now();")
w('')

for at in atividades:
    w("insert into public.carbon_atividades_campo")
    w("  (projeto_id, grupo_id, relatorio, inicio, termino, atividade, instituicao,")
    w("   tipo, linha_estrategica, evidencia, valor, status, observacoes, origem_aba, origem_linha)")
    w("select p.id,")
    if at['grupo']:
        w("  (select g.id from public.carbon_grupos_comunitarios g where g.projeto_id = p.id and g.chave = %s)," % sql(at['grupo']))
    else:
        w("  null,")
    w("  %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %d" % (
        sql(at['relatorio']), sql(at['inicio']), sql(at['termino']), sql(at['atividade']),
        sql(at['instituicao']), sql(at['tipo']), sql(at['linha']), sql(at['evidencia']),
        sql(at['valor']), sql(at['status']), sql(at['obs']), sql(at['aba']), at['linha_origem']))
    w("  from public.carbon_projetos p")
    w("  where p.nome ilike '%parakan%' or p.nome ilike '%awaet%'")
    w("on conflict on constraint carbon_ativ_campo_origem_uq do update set")
    w("  atividade = excluded.atividade, status = excluded.status,")
    w("  evidencia = excluded.evidencia, observacoes = excluded.observacoes,")
    w("  atualizado_em = now();")
w('')
w('commit;')

SAIDA.write_text('\n'.join(L) + '\n', encoding='utf-8', newline='\n')

# ===== Relatorio =============================================================

print('Escrito: %s' % SAIDA)
print('  %d aldeias | %d eixos | %d atividades de campo' % (len(aldeias), len(eixos), len(atividades)))
for (grupo, cc), c in CICLOS.items():
    ants = sum(a['valor'] for a in antecipacoes if (a['grupo'], a['ciclo']) == (grupo, cc))
    lans = [l for l in lancamentos if (l['grupo'], l['ciclo']) == (grupo, cc)]
    semc = sum(1 for l in lans if l['tem_comprovante'] is False)
    comc = sum(1 for l in lans if l['tem_comprovante'] is True)
    cps = [p for p in comprovantes if (p['grupo'], p['ciclo']) == (grupo, cc)]
    print('  %-6s %-26s antecipado %12.2f | declarado %13.2f (%4d linhas: %d com, %d sem, %d n/i) | %4d comprov (%.2f)' % (
        grupo, c['nome'], ants, sum(l['valor'] for l in lans), len(lans),
        comc, semc, len(lans) - comc - semc, len(cps), sum(p['valor'] for p in cps)))
print('')
print('Atividades por relatorio:', dict(Counter(a['relatorio'] for a in atividades)))
print('Descricoes retidas por LGPD (valor sempre importado):')
for k, n in retidos.most_common():
    print('  %4d  %s' % (n, k))
print('Observacoes de frase unica descartadas (nome mora nelas): %d' % obs_descartadas)
# Relatorio POR CODIGO, e nao por termo: imprimir o termo colocaria o nome de volta
# na tela e no log do terminal.
_por_codigo = Counter()
for _termo, _n in trocas_pessoa.items():
    _por_codigo[TERMOS_PESSOA[_termo]] += _n
print('Nomes de pessoa trocados por marcador: %d ocorrencias em %d pessoas distintas'
      % (sum(_por_codigo.values()), len(_por_codigo)))
print('  ' + '  '.join('%s=%d' % (c, n) for c, n in sorted(_por_codigo.items())))
print('Lancamentos conciliados automaticamente por valor e mes: %d' % conciliados)
if descartes:
    print('Linhas nao importadas: %d (sem valor ou sem data)' % len(descartes))
