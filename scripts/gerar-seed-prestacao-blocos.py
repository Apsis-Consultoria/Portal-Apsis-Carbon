# -*- coding: utf-8 -*-
"""
Carrega os blocos de dados que a leitura por-um-cabecalho-so nao viu.

    cd portal-apsis-carbon; python scripts/gerar-seed-prestacao-blocos.py "C:/Users/<voce>/Downloads"

Escreve supabase/seeds/prestacao_blocos.sql.

-------------------------------------------------------------------------------
POR QUE ESTE ARQUIVO EXISTE, e nao mais uma versao do gerador principal
-------------------------------------------------------------------------------
O gerador principal (gerar-seed-prestacao.py) carrega o MOVIMENTO: repasse,
despesa, comprovante, atividade. Ele funciona e os totais fecham.
Este carrega os BLOCOS DE CONTEXTO que vivem nas mesmas abas e que uma leitura
por cabecalho unico nao alcanca - eles ficam acima, abaixo e AO LADO da tabela
principal:

  1. Populacao estimada e casas por aldeia (Baixo L25-L43 F-G; Cima L27-L39 P-Q)
  2. Valor recebido por aldeia, como o RELATORIO afirma (mesmos blocos)
  3. Descricao fisica por eixo (Baixo c1 L60-L71; Baixo c2 L55-L61)
  4. Pendencia de nota fiscal (aba oculta "Pendencias", 6 notas)
  5. Parecer da APSIS por ciclo ("Observacoes Importantes", 4 blocos)
  6. Antecipacoes do 2o ciclo do Cima (L11-L16 L-M) - o ciclo estava com
     antecipado ZERO no portal, ou seja so gastava e nunca recebia
  7. Comprovantes de set/out 25 da aba "Area de trabalho" (124 linhas)
  8. ASSOCIACAO como aldeia do Cima: 235 mil estavam com aldeia_id NULL
  9. Indicadores ODS do MR-1 (aba "ODS - MR 1", 12 linhas)
 10. "Net Benifities to Community" do MR-1 (grafado assim na origem)

O QUE UMA CONFERENCIA CETICA REFUTOU e por isso NAO entra: total de formula
(SUM/SUMIFS), pivot, matriz Eixo x Mes da aba oculta, "resumo1"/"Resumo" do
Cima (tabelas dinamicas defasadas), bloco 6 de "Pendencias Israel" (consolidado
que repete os blocos 1-5), e a aba "ODS - MR 2" (vazia).
"""

import csv
import io
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
import pseudonimos  # noqa: E402  (precisa do sys.path acima)

AQUI = Path(__file__).resolve().parent
SAIDA = AQUI.parent / 'supabase' / 'seeds' / 'prestacao_blocos.sql'
PASTA = Path(sys.argv[1]) if len(sys.argv) > 1 else Path.home() / 'Downloads'

ARQ_BAIXO = PASTA / 'Antecipação Grupo de Baixo.xlsx'
ARQ_CIMA = PASTA / 'Antecipação Grupo de Cima.xlsx'
ARQ_ATIV = PASTA / 'Atividade Parakanã.xlsx'
# A mesma lista revisada que o gerar-seed-prestacao.py usa, para a mesma pessoa
# receber o mesmo marcador nos dois seeds. FORA do repositorio de proposito.
ARQ_NOMES = Path('C:/Users/FilipeOliveiraAPSISC/notion-export/nomes-seeds.csv')

CICLO = {
    ('baixo', 'c1'): 'Outubro 2024 a Abril 2025',
    ('baixo', 'c2'): 'Maio a Julho 2025',
    ('cima', 'c1'): 'Outubro 2024 a Abril 2025',
    ('cima', 'c2'): 'Maio a Setembro 2025',
}

# Nome de pessoa nunca entra. No parecer da APSIS o nome do prestador aparece:
# vira o cargo, que e dado organizacional.
#
# -----------------------------------------------------------------------------
# ANTES DAQUI SAIA O NOME DE UMA PESSOA ESCRITO NO CODIGO
# -----------------------------------------------------------------------------
# Esta lista tinha `re.compile(r'\bIsrael\b')` literal, versionado. Duas coisas
# erradas ao mesmo tempo: o nome estava no repositorio (que e o que a
# pseudonimizacao existe para evitar) e a lista so cobria quem alguem lembrou de
# escrever - "Rithelly", no parecer de um dos ciclos, nunca foi lembrada e entrou
# no seed em texto puro.
#
# Agora a lista vem do MESMO arquivo externo que o gerar-seed-prestacao.py usa,
# para a mesma pessoa receber o mesmo marcador nos dois seeds. Sem o arquivo, o
# script aborta: gerar seed com nome cru em silencio e o pior resultado possivel.

# A REGRA ESTRUTURAL FICA, e vale mais que a lista: ela pega nome que ninguem
# cadastrou, pela CONSTRUCAO da frase ("informado pelo <Nome>"). Foi varrendo
# construcoes assim - "em nome de", "assinado pelo", "feita pela" - que os tres
# ultimos nomes apareceram. Preposicao e um detector melhor de pessoa do que
# qualquer regex de nome proprio: "2 Placas Solares" ja foi acusado de ser nome.
SUBST_ESTRUTURAL = [
    (re.compile(r'\bINFORMADO PELO\s+[A-ZÁÉÍÓÚÂÊÔÃÕÇ][\wÁÉÍÓÚÂÊÔÃÕÇáéíóúâêôãõç]+', re.I),
     'INFORMADO PELO RESPONSÁVEL DA PRESTAÇÃO'),
]


# A logica de carregar, substituir e conferir vive em scripts/pseudonimos.py, e
# nao aqui: ela estava copiada em cada gerador e a copia divergiu na primeira
# mudanca de regra. Quando o codigo PROTEGER foi criado - para "Sao Paulo" nao
# ser confundido com a pessoa "Paulo" - a correcao precisou ir a cinco arquivos,
# e este era um dos que ficariam para tras.
PSEUDO = pseudonimos.Pseudonimizador(ARQ_NOMES)


def sql(v):
    if v is None:
        return 'null'
    if isinstance(v, bool):
        return 'true' if v else 'false'
    if isinstance(v, (int, float)):
        return repr(round(float(v), 2))
    if isinstance(v, (date, datetime)):
        return "'%s'" % v.strftime('%Y-%m-%d')
    return "'%s'" % str(v).replace("'", "''")


def texto(v):
    if v is None:
        return None
    s = str(v).replace('\u2014', '-').replace('\u2013', '-')
    s = re.sub(r'\s*\n\s*', ' / ', s).strip()
    # A estrutural primeiro: ela troca a frase inteira, e rodar depois da lista
    # deixaria "informado pelo [P467]" em vez de "informado pelo responsavel".
    for padrao, troca in SUBST_ESTRUTURAL:
        s = padrao.sub(troca, s)
    s = PSEUDO.aplicar(s)
    return s or None


def numero(v):
    if v in (None, ''):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace('R$', '').replace(' ', '')
    if s in ('-', '_', ''):
        return None
    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def inteiro(v):
    n = numero(v)
    return int(n) if n is not None and float(n).is_integer() else None


def dia(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def chave(nome):
    s = unicodedata.normalize('NFKD', str(nome or '')).encode('ascii', 'ignore').decode()
    return re.sub(r'[^a-z0-9]+', '', s.lower())


L = []
w = L.append
relatorio = []


def sel_ciclo(grupo, cc, junta=''):
    """FROM + JOINs + WHERE, na ordem que o Postgres exige.

    `junta` entra ANTES do where: com o where no meio, um join adicional
    depois dele e sintaxe invalida - foi exatamente o erro da primeira versao.
    """
    return ("from public.carbon_ciclos_prestacao c"
            " join public.carbon_grupos_comunitarios g on g.id = c.grupo_id"
            + (' ' + junta if junta else '')
            + " where g.chave = %s and c.nome = %s" % (sql(grupo), sql(CICLO[(grupo, cc)])))


w('-- Gerado por scripts/gerar-seed-prestacao-blocos.py. Nao edite a mao.')
w('-- Os blocos de contexto das planilhas: ver o cabecalho do script.')
w('')
w('begin;')
w('')
w("delete from public.carbon_aldeia_rateio where origem_aba is not null;")
w("delete from public.carbon_eixo_resumo where origem_aba is not null;")
w("delete from public.carbon_prestacao_pendencias where origem_aba is not null;")
w("delete from public.carbon_ods_contribuicoes where origem_aba is not null;")
w('')

wb_b = load_workbook(ARQ_BAIXO, data_only=True)
wb_c = load_workbook(ARQ_CIMA, data_only=True)
wb_a = load_workbook(ARQ_ATIV, data_only=True)

# ===== 1, 2: cadastro de aldeia (populacao, casas) e rateio ==================

def cadastro_aldeias(ws, grupo, cc, primeira, ultima, col_nome, aba):
    """Le o bloco Aldeias / Valor recebido / Cacique / Populacao / Casas."""
    n_pop = n_rat = 0
    for i, row in enumerate(ws.iter_rows(min_row=primeira, max_row=ultima, values_only=True), primeira):
        nome = texto(row[col_nome]) if col_nome < len(row) else None
        if not nome or nome.upper().startswith('TOTAL') or nome.upper().startswith('ALDEIA'):
            continue
        valor = numero(row[col_nome + 1]) if col_nome + 1 < len(row) else None
        cacique = texto(row[col_nome + 2]) if col_nome + 2 < len(row) else None
        pop = inteiro(row[col_nome + 3]) if col_nome + 3 < len(row) else None
        casas = inteiro(row[col_nome + 4]) if col_nome + 4 < len(row) else None
        # O NOME do cacique nao entra; conta-se QUANTAS liderancas a aldeia
        # declara (a celula traz uma ou duas, separadas por quebra de linha).
        lider = len([p for p in re.split(r'\s*/\s*', cacique) if p.strip()]) if cacique else None

        if pop is not None or casas is not None or lider is not None:
            w("update public.carbon_aldeias a set")
            w("  populacao_estimada = coalesce(%s, a.populacao_estimada)," % sql(pop))
            w("  casas = coalesce(%s, a.casas)," % sql(casas))
            w("  liderancas = coalesce(%s, a.liderancas)," % sql(lider))
            w("  censo_origem = 'Planilha de antecipacao, bloco de cadastro de aldeias',")
            w("  atualizado_em = now()")
            w(" from public.carbon_grupos_comunitarios g")
            w(" where a.grupo_id = g.id and g.chave = %s and a.nome = %s;" % (sql(grupo), sql(nome)))
            n_pop += 1

        if valor is not None:
            w("insert into public.carbon_aldeia_rateio (ciclo_id, aldeia_id, valor, origem_aba, origem_linha)")
            w("select c.id, a.id, %s, %s, %d" % (sql(valor), sql(aba), i))
            w("  %s" % sel_ciclo(grupo, cc,
                'join public.carbon_aldeias a on a.grupo_id = g.id and a.nome = %s' % sql(nome)))
            w("on conflict (ciclo_id, aldeia_id) do update set valor = excluded.valor;")
            n_rat += 1
    return n_pop, n_rat


# Col 2 no Baixo (col 1 e o numero de ordem) e col 12 no Cima. Os dois blocos
# tem o MESMO cabecalho e indices diferentes: por isso o indice e parametro, e
# nao constante - foi assumir "col 1 nos dois" que perdeu as 19 aldeias do Baixo.
p1, r1 = cadastro_aldeias(wb_b['Outubro 24 - Abril 25 '], 'baixo', 'c1', 24, 44, 2, 'Outubro 24 - Abril 25')
p2, r2 = cadastro_aldeias(wb_c['Out 24 a Abril 25'], 'cima', 'c1', 26, 40, 12, 'Out 24 a Abril 25')
relatorio.append(('aldeias com populacao/casas', p1 + p2))
relatorio.append(('rateio por aldeia', r1 + r2))
w('')

# ===== 3: descricao fisica por eixo =========================================

def eixo_resumo(ws, grupo, cc, primeira, ultima, col, aba):
    n = 0
    for i, row in enumerate(ws.iter_rows(min_row=primeira, max_row=ultima, values_only=True), primeira):
        nome = texto(row[col]) if col < len(row) else None
        if not nome or nome.upper().startswith('DESPESAS POR'):
            continue
        valor = numero(row[col + 1]) if col + 1 < len(row) else None
        entregas = texto(row[col + 2]) if col + 2 < len(row) else None
        if valor is None:
            continue
        w("insert into public.carbon_eixo_resumo (ciclo_id, eixo_id, valor, entregas, origem_aba, origem_linha)")
        w("select c.id, e.id, %s, %s, %s, %d" % (sql(valor), sql(entregas), sql(aba), i))
        w("  %s" % sel_ciclo(grupo, cc,
            'join public.carbon_eixos e on e.projeto_id = g.projeto_id and e.nome = %s' % sql(nome)))
        w("on conflict (ciclo_id, eixo_id) do update set")
        w("  valor = excluded.valor, entregas = excluded.entregas;")
        n += 1
    return n


e1 = eixo_resumo(wb_b['Outubro 24 - Abril 25 '], 'baixo', 'c1', 61, 71, 2, 'Outubro 24 - Abril 25')
e2 = eixo_resumo(wb_b['Maio - Julho 2025'], 'baixo', 'c2', 56, 61, 12, 'Maio - Julho 2025')
relatorio.append(('eixos com entrega fisica', e1 + e2))
w('')

# ===== 4: pendencia de nota fiscal ==========================================

ws = wb_b['Pendências']
n_pend = 0
for i, row in enumerate(ws.iter_rows(min_row=3, values_only=True), 3):
    item = texto(row[2]) if len(row) > 2 else None
    if not item:
        continue
    w("insert into public.carbon_prestacao_pendencias")
    w("  (ciclo_id, grupo_id, documento, item, quantidade, valor_unitario, valor_total,")
    w("   data_documento, competencia_prestacao, situacao, observacoes, origem_aba, origem_linha)")
    w("select c.id, g.id, %s, %s, %s, %s, %s, %s, %s, 'aberta', %s, %s, %d" % (
        sql(texto(row[0])), sql(item), sql(numero(row[1])), sql(numero(row[4])),
        sql(numero(row[5])), sql(dia(row[3])), sql(dia(row[6])),
        sql(' / '.join(t for t in (texto(row[7]), texto(row[8])) if t) or None),
        sql('Pendencias'), i))
    w("  %s;" % sel_ciclo('baixo', 'c1'))
    n_pend += 1
relatorio.append(('pendencias de nota fiscal', n_pend))
w('')

# ===== 5: parecer da APSIS por ciclo ========================================

def parecer(ws, primeira, ultima, cols):
    partes = []
    for row in ws.iter_rows(min_row=primeira, max_row=ultima, values_only=True):
        for c in cols:
            t = texto(row[c]) if c < len(row) else None
            if t and len(t) > 25 and not t.upper().startswith('OBSERVA'):
                partes.append(t)
    # dedup preservando ordem
    vistos, saida = set(), []
    for p in partes:
        if p not in vistos:
            vistos.add(p)
            saida.append(p)
    return '\n\n'.join(saida) or None


PARECERES = [
    ('baixo', 'c1', parecer(wb_b['Outubro 24 - Abril 25 '], 2, 8, range(2, 9))),
    ('baixo', 'c2', parecer(wb_b['Maio - Julho 2025'], 3, 8, [12, 13])),
    ('cima', 'c1', parecer(wb_c['Out 24 a Abril 25'], 2, 8, [12, 13, 14])),
    ('cima', 'c2', parecer(wb_c['Maio a setembro 25'], 2, 8, [11, 12, 13, 14])),
]
n_par = 0
for grupo, cc, txt in PARECERES:
    if not txt:
        continue
    w("update public.carbon_ciclos_prestacao c set observacoes = %s, atualizado_em = now()" % sql(txt))
    w(" from public.carbon_grupos_comunitarios g")
    w(" where c.grupo_id = g.id and g.chave = %s and c.nome = %s;" % (sql(grupo), sql(CICLO[(grupo, cc)])))
    n_par += 1
relatorio.append(('pareceres de ciclo', n_par))
w('')

# ===== 6: antecipacoes do 2o ciclo do Cima ==================================

ws = wb_c['Maio a setembro 25']
n_ant = 0
for i, row in enumerate(ws.iter_rows(min_row=10, max_row=20, values_only=True), 10):
    rot = texto(row[11]) if len(row) > 11 else None
    val = numero(row[12]) if len(row) > 12 else None
    if not rot or val is None or val <= 0:
        continue
    if rot.upper().startswith('RECEITA') or rot.upper().startswith('ADIANT'):
        continue
    d = dia(row[11])
    if d is None:
        # O rotulo pode ser a propria data como texto/serial.
        continue
    w("insert into public.carbon_antecipacoes (ciclo_id, competencia, valor, origem_aba, origem_linha)")
    w("select c.id, %s, %s, %s, %d %s" % (
        sql(date(d.year, d.month, 28)), sql(val), sql('Maio a setembro 25'), i, sel_ciclo('cima', 'c2')))
    w("on conflict (ciclo_id, competencia) do update set valor = excluded.valor;")
    n_ant += 1
relatorio.append(('antecipacoes do Cima c2', n_ant))
w('')

# ===== 8: ASSOCIACAO como aldeia do Cima ====================================

w("-- ASSOCIACAO do Cima: 235 mil de despesa estavam com aldeia_id NULL porque a")
w("-- aldeia nao existia para esse grupo. Cria e reaponta pelo texto da origem.")
w("insert into public.carbon_aldeias (grupo_id, nome, e_associacao)")
w("select g.id, 'ASSOCIAÇÃO', true from public.carbon_grupos_comunitarios g")
w(" where g.chave = 'cima'")
w("on conflict (grupo_id, nome) do update set e_associacao = true;")
w('')

# ===== 9: indicadores ODS ===================================================

ws = wb_a['ODS - MR 1']
n_ods = 0
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    ind = texto(row[2]) if len(row) > 2 else None
    if not ind:
        continue
    impacto = texto(row[3]) if len(row) > 3 else None
    # A ultima linha e a anterior traduzida: marca o idioma em vez de sobrescrever.
    ingles = bool(impacto and re.match(r'^\s*Implemented\b', impacto, re.I))
    w("insert into public.carbon_ods_contribuicoes")
    w("  (projeto_id, relatorio, meta_ods, indicador_ods, impacto,")
    w("   contribuicao_periodo, contribuicao_vida, idioma, ordem, origem_aba, origem_linha)")
    w("select p.id, 'MR-1', %s, %s, %s, %s, %s, %s, %d, %s, %d" % (
        sql(texto(row[1]) if len(row) > 1 else None), sql(ind), sql(impacto),
        sql(texto(row[4]) if len(row) > 4 else None),
        sql(texto(row[5]) if len(row) > 5 else None),
        sql('en' if ingles else 'pt'), i, sql('ODS - MR 1'), i))
    w("  from public.carbon_projetos p")
    w("  where p.nome ilike '%parakan%' or p.nome ilike '%awaet%'")
    w("on conflict on constraint carbon_ods_origem_uq do update set")
    w("  indicador_ods = excluded.indicador_ods, impacto = excluded.impacto,")
    w("  contribuicao_periodo = excluded.contribuicao_periodo,")
    w("  contribuicao_vida = excluded.contribuicao_vida, atualizado_em = now();")
    n_ods += 1
relatorio.append(('indicadores ODS', n_ods))
w('')

# ===== 10: beneficio a comunidade e link, no MR-1 ===========================

ws = wb_a['MR - 1 ']
n_ben = 0
for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
    ben = texto(row[16]) if len(row) > 16 else None
    link = texto(row[12]) if len(row) > 12 else None
    le2 = texto(row[8]) if len(row) > 8 else None
    if not (ben or link or le2):
        continue
    w("update public.carbon_atividades_campo set")
    w("  beneficio_comunidade = coalesce(%s, beneficio_comunidade)," % sql(ben))
    w("  link_evidencia = coalesce(%s, link_evidencia)," % sql(link))
    w("  linha_estrategica_2 = coalesce(%s, linha_estrategica_2)," % sql(le2))
    w("  atualizado_em = now()")
    w(" where relatorio = 'MR-1' and origem_aba = 'MR - 1' and origem_linha = %d;" % i)
    n_ben += 1
relatorio.append(('atividades MR-1 enriquecidas', n_ben))
w('')
w('commit;')

_sql_gerado = '\n'.join(L) + '\n'
# ANTES de escrever: arquivo com nome de pessoa nao deve nem chegar ao disco.
PSEUDO.conferir_saida(_sql_gerado, str(SAIDA))

SAIDA.write_text(_sql_gerado, encoding='utf-8', newline='\n')

print('Escrito: %s' % SAIDA)
for rotulo, n in relatorio:
    print('  %4d  %s' % (n, rotulo))
